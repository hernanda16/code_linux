import itertools
from openpyxl import Workbook

num_groups = 46
num_combinations = 2  # 2C46 (2 Combination of 46)

# Generate all combinations of groups meeting each other
all_combinations = list(itertools.combinations(
    range(1, num_groups + 1), num_combinations))

# Create a new workbook
workbook = Workbook()
sheet = workbook.active

# Print the sessions and save in Excel
for session in range(45):
    session_num = session + 1
    row_num = session + 1
    sheet.cell(row=row_num, column=1).value = f"Session {session_num}"

    temp_list = []
    for n in range(46):
        for inner_tuple in all_combinations:
            if inner_tuple[0] == n and not any(item in temp_list for item in inner_tuple):
                temp_list.append(inner_tuple[1])
                column_num = n + 2  # Start from column 2
                sheet.cell(row=row_num, column=column_num).value = str(
                    inner_tuple)
                all_combinations.remove(inner_tuple)
                break

# Save the workbook
workbook.save("sessions.xlsx")
